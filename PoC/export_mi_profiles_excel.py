# export_mi_profiles_excel.py
# python export_mi_profiles_excel.py --root output --out mi_profiles_summary.xlsx
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import Counter

from rdflib import Graph, URIRef
from rdflib.namespace import RDF, RDFS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ====== Ajuste se seu namespace mudar ======
ONTO_NS = "https://techcoop.com.br/ontomi#"
DCTERMS_NS = "http://purl.org/dc/terms/"
IAO_REFERS_TO = URIRef("http://purl.obolibrary.org/obo/IAO_0000136")  # refersTo/isAbout


# Ordem canônica do seu vetor (0..7)
INTELLIGENCES = [
    "Logical-Mathematical",
    "Spatial",
    "Linguistic",
    "Musical",
    "Bodily-Kinesthetic",
    "Interpersonal",
    "Intrapersonal",
    "Naturalist",
]

VECTOR_SCORE_PRED_BY_MI = {
    "Logical-Mathematical": URIRef(ONTO_NS + "hasLogicalMathematicalScore"),
    "Spatial":              URIRef(ONTO_NS + "hasSpatialScore"),
    "Linguistic":           URIRef(ONTO_NS + "hasLinguisticScore"),
    "Musical":              URIRef(ONTO_NS + "hasMusicalScore"),
    "Bodily-Kinesthetic":   URIRef(ONTO_NS + "hasBodilyKinestheticScore"),
    "Interpersonal":        URIRef(ONTO_NS + "hasInterpersonalScore"),
    "Intrapersonal":        URIRef(ONTO_NS + "hasIntrapersonalScore"),
    "Naturalist":           URIRef(ONTO_NS + "hasNaturalistScore"),
}

P_MI_VECTOR = URIRef(ONTO_NS + "miVector")
P_HAS_PROFILE_VECTOR = URIRef(ONTO_NS + "hasProfileVector")
P_HAS_PART = URIRef(DCTERMS_NS + "hasPart")

P_HAS_PRIMARY_ACT = URIRef(ONTO_NS + "hasPrimaryActivation")
P_HAS_ACT = URIRef(ONTO_NS + "hasActivation")

P_ACT_TYPE = URIRef(ONTO_NS + "hasActivationType")
P_ACT_SCORE = URIRef(ONTO_NS + "hasActivationScore")

ACT_PRIMARY = URIRef(ONTO_NS + "Primary")
ACT_SECONDARY = URIRef(ONTO_NS + "Secondary")

T_EXPLANATION_DOCUMENT = URIRef(ONTO_NS + "ExplanationDocument")


def _as_float(x) -> Optional[float]:
    try:
        return float(str(x))
    except Exception:
        return None


def _doc_slug_from_uri(doc_uri: URIRef) -> str:
    s = str(doc_uri)
    if s.startswith("urn:doc:"):
        return s.replace("urn:doc:", "")
    return s.rsplit("/", 1)[-1]


def _label_or_local(g: Graph, uri: URIRef) -> str:
    """rdfs:label -> localname."""
    lab = next(g.objects(uri, RDFS.label), None)
    if lab:
        return str(lab)
    s = str(uri)
    if "#" in s:
        return s.split("#")[-1]
    return s.rsplit("/", 1)[-1]


def _norm_key(s: str) -> str:
    """normaliza pra mapear label/localname -> canônico (remove pontuação e caixa)."""
    s = (s or "").strip().lower()
    out = []
    for ch in s:
        if ch.isalnum():
            out.append(ch)
    return "".join(out)


# mapeia diferentes labels/localnames possíveis para o canônico do vetor
_CANON_MAP: Dict[str, str] = {}
for mi in INTELLIGENCES:
    _CANON_MAP[_norm_key(mi)] = mi
# tolerâncias comuns
_CANON_MAP[_norm_key("LogicalMathematical")] = "Logical-Mathematical"
_CANON_MAP[_norm_key("BodilyKinesthetic")] = "Bodily-Kinesthetic"


def canon_mi(label_or_local: str) -> Optional[str]:
    return _CANON_MAP.get(_norm_key(label_or_local))


def find_profile_files(root: Path) -> List[Path]:
    return sorted([p for p in root.rglob("instances_fragments_profile.ttl") if p.is_file()])


def parse_one_ttl(ttl_path: Path) -> Tuple[List[dict], List[dict]]:
    g = Graph()
    g.parse(ttl_path.as_posix(), format="turtle")

    docs = list(g.subjects(RDF.type, T_EXPLANATION_DOCUMENT))
    if not docs:
        # fallback: urn:doc:
        docs = [URIRef(s) for s in set(str(x) for x in g.subjects()) if str(s).startswith("urn:doc:")]

    doc_rows: List[dict] = []
    frag_rows: List[dict] = []

    for doc in docs:
        # --- vetor do documento (MIProfileVector)
        vec = next(g.objects(doc, P_HAS_PROFILE_VECTOR), None)
        mi_vector_str = next(g.objects(vec, P_MI_VECTOR), None) if vec else None

        vec_scores: Dict[str, Optional[float]] = {mi: None for mi in INTELLIGENCES}
        if vec:
            for mi in INTELLIGENCES:
                val = next(g.objects(vec, VECTOR_SCORE_PRED_BY_MI[mi]), None)
                vec_scores[mi] = _as_float(val) if val is not None else None

        # primária pelo vetor
        primary_by_vector = ""
        primary_by_vector_score = None
        numeric = [(mi, v) for mi, v in vec_scores.items() if isinstance(v, (int, float))]
        if numeric:
            mi_max, v_max = max(numeric, key=lambda t: t[1])
            if v_max is not None and float(v_max) > 0.0:
                primary_by_vector = mi_max
                primary_by_vector_score = float(v_max)

        # --- contadores por doc (prim/sec a partir dos FRAGMENTOS)
        prim_counts = Counter({mi: 0 for mi in INTELLIGENCES})
        sec_counts = Counter({mi: 0 for mi in INTELLIGENCES})

        fragments = list(g.objects(doc, P_HAS_PART))

        for frag in fragments:
            frag_label = next(g.objects(frag, RDFS.label), None)
            frag_label = str(frag_label) if frag_label else ""

            # por fragmento, vamos calcular max score por MI (prim e sec)
            frag_primary_scores: Dict[str, Optional[float]] = {mi: None for mi in INTELLIGENCES}
            frag_secondary_scores: Dict[str, Optional[float]] = {mi: None for mi in INTELLIGENCES}

            # primárias preferindo hasPrimaryActivation
            primary_acts = list(g.objects(frag, P_HAS_PRIMARY_ACT))

            # fallback: se vazio, usa hasActivationType=Primary
            if not primary_acts:
                for act in g.objects(frag, P_HAS_ACT):
                    if next(g.objects(act, P_ACT_TYPE), None) == ACT_PRIMARY:
                        primary_acts.append(act)

            for act in primary_acts:
                intel_uri = next(g.objects(act, IAO_REFERS_TO), None)
                if not isinstance(intel_uri, URIRef):
                    continue
                intel_raw = _label_or_local(g, intel_uri)
                mi = canon_mi(intel_raw)
                if not mi:
                    continue
                score = _as_float(next(g.objects(act, P_ACT_SCORE), None))
                # marca presença (contagem por fragmento: conta 1 por MI se apareceu como primária)
                # (se tiver múltiplas primárias iguais no mesmo frag, evita somar duplicado depois)
                prev = frag_primary_scores[mi]
                if score is not None:
                    frag_primary_scores[mi] = score if prev is None else max(prev, score)
                else:
                    frag_primary_scores[mi] = prev if prev is not None else 0.0

            # secundárias: hasActivationType=Secondary
            for act in g.objects(frag, P_HAS_ACT):
                if next(g.objects(act, P_ACT_TYPE), None) != ACT_SECONDARY:
                    continue
                intel_uri = next(g.objects(act, IAO_REFERS_TO), None)
                if not isinstance(intel_uri, URIRef):
                    continue
                intel_raw = _label_or_local(g, intel_uri)
                mi = canon_mi(intel_raw)
                if not mi:
                    continue
                score = _as_float(next(g.objects(act, P_ACT_SCORE), None))
                prev = frag_secondary_scores[mi]
                if score is not None:
                    frag_secondary_scores[mi] = score if prev is None else max(prev, score)
                else:
                    frag_secondary_scores[mi] = prev if prev is not None else 0.0

            # atualiza contagens por doc (1 por fragmento onde MI apareceu)
            for mi in INTELLIGENCES:
                if frag_primary_scores[mi] is not None:
                    prim_counts[mi] += 1
                if frag_secondary_scores[mi] is not None:
                    sec_counts[mi] += 1

            # linha por fragmento com colunas por MI (scores)
            frag_row = {
                "ttl_path": ttl_path.as_posix(),
                "doc_id": str(doc),
                "doc_slug": _doc_slug_from_uri(doc),
                "fragment_uri": str(frag),
                "fragment_label": frag_label,
            }
            for mi in INTELLIGENCES:
                frag_row[f"Primary_{mi}_score"] = frag_primary_scores[mi]
            for mi in INTELLIGENCES:
                frag_row[f"Secondary_{mi}_score"] = frag_secondary_scores[mi]
            frag_rows.append(frag_row)

        # linha do doc com colunas por MI (counts)
        doc_row = {
            "ttl_path": ttl_path.as_posix(),
            "doc_id": str(doc),
            "doc_slug": _doc_slug_from_uri(doc),
            "miVector": str(mi_vector_str) if mi_vector_str is not None else "",
            "primary_by_vector": primary_by_vector,
            "primary_by_vector_score": primary_by_vector_score,
        }
        for mi in INTELLIGENCES:
            doc_row[f"{mi}_score"] = vec_scores.get(mi)
        for mi in INTELLIGENCES:
            doc_row[f"Primary_{mi}_count"] = prim_counts.get(mi, 0)
        for mi in INTELLIGENCES:
            doc_row[f"Secondary_{mi}_count"] = sec_counts.get(mi, 0)

        doc_rows.append(doc_row)

    return doc_rows, frag_rows


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def write_excel(out_xlsx: Path, doc_rows: List[dict], frag_rows: List[dict]) -> None:
    wb = Workbook()

    # ---------- documents ----------
    ws = wb.active
    ws.title = "documents"

    doc_header = [
        "ttl_path", "doc_id", "doc_slug",
        "miVector", "primary_by_vector", "primary_by_vector_score",
    ]
    # scores do MIProfileVector na ordem do seu vetor
    for mi in INTELLIGENCES:
        doc_header.append(f"{mi}_score")
    # uma coluna por MI (prim/sec counts)
    for mi in INTELLIGENCES:
        doc_header.append(f"Primary_{mi}_count")
    for mi in INTELLIGENCES:
        doc_header.append(f"Secondary_{mi}_count")

    ws.append(doc_header)
    _style_header(ws, len(doc_header))

    for r in doc_rows:
        ws.append([r.get(h, "") for h in doc_header])

    # larguras básicas
    widths = {"A": 45, "B": 28, "C": 18, "D": 40, "E": 22, "F": 22}
    for col in range(1, len(doc_header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(get_column_letter(col), 18)

    # ---------- fragments ----------
    ws2 = wb.create_sheet("fragments")
    frag_header = [
        "ttl_path", "doc_id", "doc_slug",
        "fragment_uri", "fragment_label",
    ]
    # uma coluna por MI (prim/sec scores)
    for mi in INTELLIGENCES:
        frag_header.append(f"Primary_{mi}_score")
    for mi in INTELLIGENCES:
        frag_header.append(f"Secondary_{mi}_score")

    ws2.append(frag_header)
    _style_header(ws2, len(frag_header))

    for r in frag_rows:
        ws2.append([r.get(h, "") for h in frag_header])

    widths2 = {"A": 45, "B": 28, "C": 18, "D": 45, "E": 45}
    for col in range(1, len(frag_header) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = widths2.get(get_column_letter(col), 18)

    wb.save(out_xlsx.as_posix())


def main() -> int:
    ap = argparse.ArgumentParser(description="Exporta MI primárias/secundárias (1 MI por coluna) a partir de instances_fragments_profile.ttl.")
    ap.add_argument("--root", default="output", help="Pasta raiz onde estão as subpastas com TTLs.")
    ap.add_argument("--out", default="mi_profiles_summary.xlsx", help="Caminho do Excel de saída.")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    out_xlsx = Path(args.out).resolve()

    if not root.exists():
        print(f"[ERRO] root não existe: {root}")
        return 1

    files = find_profile_files(root)
    if not files:
        print(f"[ERRO] Nenhum instances_fragments_profile.ttl encontrado em: {root}")
        return 2

    all_doc_rows: List[dict] = []
    all_frag_rows: List[dict] = []

    for ttl in files:
        try:
            doc_rows, frag_rows = parse_one_ttl(ttl)
            all_doc_rows.extend(doc_rows)
            all_frag_rows.extend(frag_rows)
            print(f"[OK] {ttl} | docs={len(doc_rows)} | frags={len(frag_rows)}")
        except Exception as e:
            print(f"[FALHOU] {ttl}: {e!r}")

    write_excel(out_xlsx, all_doc_rows, all_frag_rows)
    print(f"\n[XLSX OK] -> {out_xlsx}")
    print(f"Docs: {len(all_doc_rows)} | Fragments: {len(all_frag_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
