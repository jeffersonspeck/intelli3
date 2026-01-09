"""\
PT-BR: Orquestrador do pipeline Intelli3 para execução em lote.
EN: Intelli3 pipeline orchestrator for batch execution.

run_batch.py — Orquestrador do pipeline Intelli3 (execução em lote).

Este script é o ponto de entrada do Intelli3 para processamento em lote de
documentos textuais (.txt), executando de forma controlada e incremental
as etapas S1 a S7 do pipeline.

Visão geral do funcionamento:

1. Descoberta de entradas
   - Lê todos os arquivos .txt presentes na pasta de origem (--source-dir).
   - Para cada arquivo, gera um identificador (slug) seguro baseado no nome.

2. Organização da saída
   - Para cada documento de entrada, cria uma pasta dedicada em:
       ./output/<doc_slug>/
   - Todos os artefatos gerados pelas etapas S1–S7 são armazenados dentro
     dessa pasta.
   - Um relatório agregado do lote é escrito em:
       ./output/batch_report.json

3. Execução do pipeline (S1 → S7)

   As etapas são executadas de forma sequencial e dependente:

   S1 — Ingestão e pré-processamento
     - Limpeza textual (ex.: ftfy, clean_text, pdf_breaks).
     - Segmentação em parágrafos/fragments.
     - Detecção de idioma (LID).
     - Identificação heurística de título/resumo.
     - Gera: s1_output.json

   S2 — Instanciação RDF
     - Converte o JSON da S1 em instâncias RDF (Documento, Fragmentos).
     - Utiliza a ontologia OntoMI (TTL).
     - Suporta diferentes modos de grafo (full, instances, instances+imports).
     - Gera: instances_fragments.ttl

   S3 — Evidências semânticas
     - Analisa fragmentos e cria evidências do tipo:
         Keyword, ContextObject, DiscursiveStrategy.
     - Pode utilizar LLM (via evidences_api.py + Ollama).
     - Opcionalmente cria links onto:evokesIntelligence.
     - Gera:
         instances_fragments_evidences.ttl
         evidences_payload.json

   S4 — Ativações (a_{k,j})
     - Calcula scores de ativação por fragmento e inteligência.
     - Aplica pesos por tipo de evidência (keyword/context/strategy).
     - Define inteligência primária e secundárias via limiar theta.
     - Gera:
         instances_fragments_activations.ttl
         scores_by_fragment.json

   S5 — Vetor de perfil (MIProfileVector)
     - Agrega ativações em vetores de Inteligências Múltiplas.
     - Pode gerar vetores por fragmento, por documento ou ambos.
     - Suporta normalização L1, L2 ou softmax (com temperatura).
     - Opcionalmente escreve a string percentual onto:miVector.
     - Gera:
         instances_fragments_profile.ttl

   S6 — Consultas de validação (CQ1, CQ2, CQ3)
     - Executa consultas SPARQL sobre o grafo final:
         CQ1: inteligências evocadas
         CQ2: elementos contribuintes por inteligência
         CQ3: inteligência primária vs secundárias
     - Gera arquivos .txt com os resultados.

   S7 — Validação SHACL
     - Valida o grafo RDF final contra as restrições SHACL da OntoMI.
     - Suporta inferência (none, rdfs, owlrl).
     - Gera:
         shacl_report.ttl
         shacl_report.txt

4. Controle de dependências entre etapas
   - Se uma etapa for desativada, todas as subsequentes são automaticamente
     desabilitadas.
   - Exemplo: se S3 for desativada, S4–S7 não serão executadas.

5. Telemetria e rastreabilidade
   - Para cada documento:
       run_log.json com tempos, métricas e status de cada etapa.
   - Para o lote completo:
       batch_report.json com resumo geral (sucessos, falhas, métricas).

6. Tratamento de falhas
   - Se um documento falhar, o erro é registrado em:
       output/<doc_slug>/s1_error.txt
   - O processamento continua para os demais arquivos.

Objetivo do script:

- Fornecer um orquestrador reproduzível, extensível e controlável
  para experimentação científica com o pipeline Intelli3.
- Permitir análises incrementais (ligar/desligar etapas).
- Garantir rastreabilidade completa dos resultados por documento
  e por execução em lote.

Este script é parte integrante da pesquisa de Mestrado em Ciência da
Computação associada ao projeto Intelli3.
"""
from __future__ import annotations
import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from rdflib import Graph, URIRef

from s1_ingest import S1Ingestor, S1Params
from s2_instantiate import instantiate_fragments_rdf, save_graph
from s3_evidences import run_s3_from_files
from s4_activations import run_s4_from_files
from s5_profile_vector import run_s5_from_files
from s6_tests import run_s6_from_files
from s7_shacl_validate import run_s7_shacl_validate_from_files

SCRIPT_DIR = Path(__file__).resolve().parent

INTELLIGENCES = [
    "Linguistic",
    "Logical-Mathematical",
    "Spatial",
    "Bodily-Kinesthetic",
    "Musical",
    "Interpersonal",
    "Intrapersonal",
    "Naturalist",
]

def slugify(name: str) -> str:
    # Eu uso esse slug para garantir nomes de pasta previsíveis.
    name = name.strip()
    name = re.sub(r"[^\w\-. ]+", "", name, flags=re.UNICODE)
    name = name.replace(" ", "_")
    return name[:120] if len(name) > 120 else name


def iter_txt_files(src_dir: Path) -> List[Path]:
    # Aqui eu filtro só os .txt para manter o lote controlado.
    return sorted([p for p in src_dir.glob("*.txt") if p.is_file()])


def _iso_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _write_json(path: Path, obj: dict) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def _parse_mivector_string(s: str) -> List[float]:
    """
    Aceita formatos como:
    "54.22,0.00,5.34,0.00,40.44,0.00,0.00,0.00"
    Retorna uma lista de float com tamanho 8 quando possível; caso contrário,
    retorna uma lista vazia.
    """
    if not s:
        return []
    parts = [p.strip() for p in str(s).split(",") if p.strip()]
    vals: List[float] = []
    for p in parts:
        p2 = p.replace("%", "").strip()
        try:
            vals.append(float(p2))
        except Exception:
            return []
    return vals


def _extract_mivector_from_profile_ttl(ttl_path: Path, vector_uri: Optional[str]) -> Optional[str]:
    """
    Lê o arquivo instances_fragments_profile.ttl e tenta extrair o valor
    onto:miVector para a URI de vetor do documento (doc_vector_uri).
    Retorna a string literal ou None.
    """
    if not ttl_path.exists() or not vector_uri:
        return None

    g = Graph()
    g.parse(ttl_path.as_posix(), format="turtle")
    subj = URIRef(vector_uri)

    for p, o in g.predicate_objects(subj):
        local = str(p)
        if local.endswith("miVector") or local.split("#")[-1] == "miVector" or local.split("/")[-1] == "miVector":
            return str(o)

    return None


def write_batch_excel(out_dir: Path, batch_rows: List[dict]) -> Path:
    """
    Cria o arquivo output/batch_summary.xlsx com 1 linha por documento:
    doc_id, doc_slug, miVector, mi_primaria, score_primario,
    colunas de score por Inteligência (MI), colunas de ativação por MI.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "batch"

    header = ["doc_id", "doc_slug", "miVector", "primary_mi", "primary_score"]
    for mi in INTELLIGENCES:
        header.append(f"{mi}_score")
    for mi in INTELLIGENCES:
        header.append(f"{mi}_active")

    ws.append(header)

    for col_idx, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    for row in batch_rows:
        doc_id = row.get("doc_id") or ""
        doc_slug = row.get("doc_slug") or ""

        profile_ttl = out_dir / doc_slug / "instances_fragments_profile.ttl"
        s5 = (row.get("steps") or {}).get("s5") or {}
        vector_uri = s5.get("doc_vector_uri")

        mi_vector_str = _extract_mivector_from_profile_ttl(profile_ttl, vector_uri) or ""
        scores = _parse_mivector_string(mi_vector_str)

        # ensure length 8
        if len(scores) != len(INTELLIGENCES):
            scores = [None] * len(INTELLIGENCES)

        # primary MI = maior score (ignorando None)
        numeric = [(i, v) for i, v in enumerate(scores) if isinstance(v, (int, float))]
        if numeric:
            primary_idx, primary_score = max(numeric, key=lambda t: t[1])
            # se tudo for 0, você pode preferir deixar vazio:
            if float(primary_score) > 0.0:
                primary_mi = INTELLIGENCES[primary_idx]
            else:
                primary_mi = ""
                primary_score = None
        else:
            primary_mi = ""
            primary_score = None

        active = [(float(v) > 0.0) if isinstance(v, (int, float)) else None for v in scores]

        line = [doc_id, doc_slug, mi_vector_str, primary_mi, primary_score]
        line.extend(scores)
        line.extend(active)
        ws.append(line)

    widths = {
        "A": 22,  # doc_id
        "B": 18,  # doc_slug
        "C": 40,  # miVector
        "D": 22,  # primary_mi
        "E": 14,  # primary_score
    }
    for col in range(1, len(header) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 18)

    xlsx_path = out_dir / "batch_summary.xlsx"
    wb.save(xlsx_path.as_posix())
    return xlsx_path


# def process_one(txt_path: Path, out_root: Path, ingestor: S1Ingestor) -> Path:
#     """Processa um TXT e retorna o caminho do s1_output.json gerado."""
def process_one(
    txt_path: Path,
    out_root: Path,
    ingestor: S1Ingestor,
    *,
    do_s2: bool,
    do_s3: bool,
    onto_path: Path,
    base_ns: Optional[str],
    s2_graph: str,
    s3_use_llm: bool,
    s3_link_evokes: bool,
    do_s4: bool,
    s4_theta: float,
    s4_w_keyword: float,
    s4_w_context: float,
    s4_w_strategy: float,
    do_s5: bool,
    s5_scope: str,
    s5_norm: str,
    s5_tau: float,
    s5_alpha_title: float,
    s5_alpha_body: float,
    s5_vec_places: int,
    s5_write_mi_vector_string: bool,  
    do_s6: bool,
    do_s7: bool,
    s7_inference: str,          
) -> dict:
    """Processa um TXT. Retorna um dicionário com métricas/telemetria."""
    name = slugify(txt_path.stem)
    out_dir = out_root / name
    out_dir.mkdir(parents=True, exist_ok=True)

    run_log: dict = {
        "doc_slug": name,
        "source_file": txt_path.name,
        "started_at": _iso_now(),
        "steps": {},
    }
    if do_s2 and (not False):
        pass

    # mede o total real do documento (S1 + S2 + S3)
    t_total0 = time.perf_counter()

    # -------------------------
    # S1
    # -------------------------
    t0 = time.perf_counter()
    doc = ingestor.ingest_file(txt_path.as_posix(), doc_id=f"urn:doc:{name}")
    dt = time.perf_counter() - t0

    s1_json = out_dir / "s1_output.json"
    # s1_json.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_json(s1_json, doc)

    # mini-resumo no console
    n_paras = len(doc.get("paragraphs", []) or [])
    # print(f"[S1 OK] {txt_path.name} -> {s1_json.as_posix()} | paragraphs={n_paras} | {dt:.3f}s")
    # return s1_json
    print(f"[S1 OK] {txt_path.name} -> {s1_json.as_posix()} | paragraphs={n_paras} | {dt:.3f}s")

    run_log["doc_id"] = doc.get("doc_id")
    run_log["steps"]["s1"] = {
        "status": "ok",
        "duration_s": round(dt, 6),
        "paragraphs": n_paras,
        "output": s1_json.name,
    }

    # -------------------------
    # S2 — instanciação RDF
    # -------------------------
    if do_s2:
        t1 = time.perf_counter()
        g, frags = instantiate_fragments_rdf(
            doc,
            onto_path=str(onto_path),
            base_instances_ns=base_ns,
            graph_mode=s2_graph,
        )
        dt2 = time.perf_counter() - t1

        ttl_out = out_dir / "instances_fragments.ttl"
        save_graph(g, ttl_out.as_posix())

        print(f"[S2 OK] {txt_path.name} -> {ttl_out.as_posix()} | fragments={len(frags)} | {dt2:.3f}s")

        run_log["steps"]["s2"] = {
            "status": "ok",
            "duration_s": round(dt2, 6),
            "fragments": len(frags),
            "onto_path": str(onto_path),
            "output": ttl_out.name,
            "base_ns": base_ns or (str(doc.get("doc_id")) + "#"),
            "graph_mode": s2_graph,
        }

    # -------------------------
    # S3 — evidências RDF
    # -------------------------
    if do_s3:
        t2 = time.perf_counter()

        s2_ttl_path = out_dir / "instances_fragments.ttl"
        s3_ttl_out = out_dir / "instances_fragments_evidences.ttl"
        s3_payload_out = out_dir / "evidences_payload.json"

        metrics = run_s3_from_files(
            s1_json_path=s1_json,
            s2_ttl_path=s2_ttl_path,
            out_ttl_path=s3_ttl_out,
            out_payload_json_path=s3_payload_out,
            onto_path=onto_path,
            use_llm=s3_use_llm,
            link_evokes_intelligence=s3_link_evokes,
        )

        dt3 = time.perf_counter() - t2

        print(
            f"[S3 OK] {txt_path.name} -> {s3_ttl_out.as_posix()} "
            f"| evidences={metrics.get('evidences_total')} "
            f"| roles={metrics.get('role_counts')} "
            f"| {dt3:.3f}s"
        )

        run_log["steps"]["s3"] = {
            "status": "ok",
            "duration_s": round(dt3, 6),
            **metrics,
        }

    # -------------------------
    # S4 — ativações RDF (a_{k,j})
    # -------------------------
    if do_s4:
        t3 = time.perf_counter()

        s3_ttl_path = out_dir / "instances_fragments_evidences.ttl"
        payload_path = out_dir / "evidences_payload.json"
        scores_out = out_dir / "scores_by_fragment.json"
        s4_ttl_out = out_dir / "instances_fragments_activations.ttl"

        metrics = run_s4_from_files(
            s3_ttl_path=s3_ttl_path,
            out_ttl_path=s4_ttl_out,
            onto_path=onto_path,
            evidences_payload_json_path=(payload_path if payload_path.exists() else None),
            scores_json_path=scores_out,
            theta=s4_theta,
            w_keyword=s4_w_keyword,
            w_context_object=s4_w_context,
            w_discursive_strategy=s4_w_strategy,
        )

        dt4 = time.perf_counter() - t3

        print(
            f"[S4 OK] {txt_path.name} -> {s4_ttl_out.as_posix()} "
            f"| activations={metrics.get('activations_total')} "
            f"| primaries={metrics.get('primary_total')} "
            f"| secondary_linked={metrics.get('secondary_linked_total')} "
            f"| {dt4:.3f}s"
        )

        run_log["steps"]["s4"] = {
            "status": "ok",
            "duration_s": round(dt4, 6),
            **metrics,
        }

    # -------------------------
    # S5 — profile vector RDF
    # -------------------------
    if do_s5:
        t5 = time.perf_counter()

        s4_ttl_path = out_dir / "instances_fragments_activations.ttl"
        s5_ttl_out = out_dir / "instances_fragments_profile.ttl"

        metrics = run_s5_from_files(
            in_ttl_path=s4_ttl_path.as_posix(),
            out_ttl_path=s5_ttl_out.as_posix(),
            s1_json_path=s1_json.as_posix(),
            scope=s5_scope,
            norm=s5_norm,  # "l1" | "l2" | "softmax"
            tau=s5_tau,
            alpha_title=s5_alpha_title,
            alpha_body=s5_alpha_body,
            vec_places=s5_vec_places,
            write_percent_string=s5_write_mi_vector_string,
        )

        dt5 = time.perf_counter() - t5

        print(
            f"[S5 OK] {txt_path.name} -> {s5_ttl_out.as_posix()} "
            f"| scope={metrics.get('scope')} norm={metrics.get('norm')} "
            f"| doc_vec={metrics.get('doc_vector_uri')} "
            f"| {dt5:.3f}s"
        )

        run_log["steps"]["s5"] = {
            "status": "ok",
            "duration_s": round(dt5, 6),
            **metrics,
        }

    # -------------------------
    # S6 — CQ1/CQ2/CQ3 (testes)
    # -------------------------
    if do_s6:
        t6 = time.perf_counter()
        s5_ttl_path = out_dir / "instances_fragments_profile.ttl"

        metrics = run_s6_from_files(
            in_ttl_path=s5_ttl_path.as_posix(),
            out_dir=out_dir.as_posix(),
        )

        dt6 = time.perf_counter() - t6
        print(
            f"[S6 OK] {txt_path.name} -> {out_dir.as_posix()} "
            f"| cq1_rows={metrics.get('cq1', {}).get('rows')} "
            f"| cq2_rows={metrics.get('cq2', {}).get('rows')} "
            f"| cq3_rows={metrics.get('cq3', {}).get('rows')} "
            f"| {dt6:.3f}s"
        )

        run_log["steps"]["s6"] = {
            "status": "ok",
            "duration_s": round(dt6, 6),
            **metrics,
        }

    # -------------------------
    # S7 — SHACL validate
    # -------------------------
    if do_s7:
        t7 = time.perf_counter()

        s5_ttl_path = out_dir / "instances_fragments_profile.ttl"
        report_ttl = out_dir / "shacl_report.ttl"
        report_txt = out_dir / "shacl_report.txt"

        metrics = run_s7_shacl_validate_from_files(
            data_path=s5_ttl_path.as_posix(),
            shacl_path=str(onto_path),
            out_report_ttl_path=report_ttl.as_posix(),
            out_report_txt_path=report_txt.as_posix(),
            inference=(None if s7_inference == "none" else s7_inference),  # pyshacl aceita None/str
            advanced=True,
            allow_infos=True,
            allow_warnings=True,
        )

        dt7 = time.perf_counter() - t7
        print(
            f"[S7 OK] {txt_path.name} -> {report_ttl.as_posix()} "
            f"| conforms={metrics.get('conforms')} results={metrics.get('results_count')} "
            f"| {dt7:.3f}s"
        )

        run_log["steps"]["s7"] = {
            "status": "ok",
            "duration_s": round(dt7, 6),
            **metrics,
        }

    run_log["ended_at"] = _iso_now()
    # total real do documento (S1 + S2 + S3)
    run_log["total_s"] = round((time.perf_counter() - t_total0), 6)

    # salva telemetria do documento
    _write_json(out_dir / "run_log.json", run_log)

    return run_log


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Roda o pipeline em lote (S1 por enquanto) para todos os .txt de uma pasta.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    # pastas
    p.add_argument("--source-dir", default="source", help="Pasta com os .txt")
    p.add_argument("--out-dir", default="output", help="Pasta de saída")

    # S2 — instanciação RDF
    p.add_argument("--no-s2", action="store_true", help="Desliga a S2 (instanciação RDF).")
    p.add_argument("--onto", default="ontomi.ttl", help="Caminho da ontologia OntoMI (TTL).")
    p.add_argument("--base-ns", default=None, help="Namespace base para instâncias (senão usa 'doc_id#').")
    p.add_argument(
        "--s2-graph",
        default="full",
        choices=["full", "instances", "instances+imports"],
        help="Modo do grafo S2: full (ontologia+instâncias), instances (só instâncias), instances+imports (instâncias + owl:imports).",
    )

    # S3 — evidências (Keyword/ContextObject/DiscursiveStrategy)
    p.add_argument("--no-s3", action="store_true", help="Desliga a S3 (evidências).")
    p.add_argument("--s3-no-llm", action="store_true", help="Tenta rodar a S3 sem LLM (se sua evidences_api suportar).")
    p.add_argument("--s3-no-evokes", action="store_true", help="Não cria onto:evokesIntelligence nas evidências.")

    # S4 — ativações por fragmento (a_{k,j})
    p.add_argument("--no-s4", action="store_true", help="Desliga a S4 (ativações).")
    p.add_argument("--theta", type=float, default=0.75, help="Limiar relativo (theta) para linkar secundárias.")
    p.add_argument("--w-keyword", type=float, default=1.00, help="Peso w_role para Keyword.")
    p.add_argument("--w-context", type=float, default=1.25, help="Peso w_role para ContextObject.")
    p.add_argument("--w-strategy", type=float, default=1.10, help="Peso w_role para DiscursiveStrategy.")

    # S5 — vetor de perfil (MIProfileVector)
    p.add_argument("--no-s5", action="store_true", help="Desliga a S5 (vetor MIProfileVector).")
    p.add_argument("--s5-scope", choices=["document", "fragment", "both"], default="both",
                help="Escopo do S5: vetor por documento, por fragmento, ou ambos.")
    p.add_argument("--s5-norm", choices=["l1", "l2", "softmax"], default="l1",
                help="Normalização do vetor (S5).")
    p.add_argument("--s5-tau", type=float, default=1.0,
                help="Temperatura do softmax (apenas se --s5-norm=softmax).")
    p.add_argument("--alpha-title", type=float, default=1.30,
                help="alpha_k para título/resumo (S5).")
    p.add_argument("--alpha-body", type=float, default=1.00,
                help="alpha_k para corpo (S5).")
    p.add_argument("--s5-vec-places", type=int, default=4,
                help="Casas decimais nos scores do vetor (S5).")
    p.add_argument("--s5-no-mi-vector-string", action="store_true",
                help="Não escreve a propriedade onto:miVector com percentuais.")

    # S6 — consultas CQ1/CQ2/CQ3 (testes)
    p.add_argument("--no-s6", action="store_true", help="Desliga a S6 (CQ1/CQ2/CQ3).")

    # S7 — validação SHACL
    p.add_argument("--no-s7", action="store_true", help="Desliga a S7 (validação SHACL).")
    p.add_argument("--s7-inference", default="rdfs", choices=["none", "rdfs", "owlrl"],
                help="Inferência usada no pySHACL (default: rdfs).")

    # S1 — opções principais
    p.add_argument("--lang-hint", default=None, choices=["pt", "en", "es"], help="Dica de idioma (restringe languages_supported)")
    p.add_argument("--cleaners", default="ftfy,clean_text,pdf_breaks", help="Lista CSV de cleaners do intelli3text")
    p.add_argument("--lid-primary", default="fasttext", help="LID primário (ex.: fasttext)")
    p.add_argument("--lid-fallback", default="none", help="LID fallback (ex.: cld3) ou 'none'")
    p.add_argument("--languages", default="pt,en,es", help="CSV de idiomas suportados (pt,en,es)")
    p.add_argument("--nlp-size", default="lg", choices=["lg", "md", "sm"], help="Preferência spaCy (lg|md|sm)")

    # segmentação
    p.add_argument("--paragraph-min-chars", type=int, default=30, help="Min chars p/ parágrafo (intelli3text)")
    p.add_argument("--lid-min-chars", type=int, default=60, help="Min chars p/ rodar LID por parágrafo (intelli3text)")

    # fallback de split quando o TXT vira 1 parágrafo gigante
    # p.add_argument("--split-mode", default="auto", choices=["auto", "sentences", "chunks", "none"], help="Fallback de split quando necessário")
    p.add_argument("--split-max-chars", type=int, default=900, help="Tamanho alvo de chunk no fallback")
    p.add_argument("--split-min-chars", type=int, default=120, help="Min chars do chunk (merge de chunks pequenos)")
    p.add_argument("--no-resegment", action="store_true", help="Desliga fallback de resegmentação")

    # título/resumo
    p.add_argument("--force-title", action="store_true", help="Força marcar algum parágrafo como título")
    p.add_argument("--title-scan-k", type=int, default=3, help="Quantos primeiros parágrafos procurar título")
    p.add_argument("--title-max-chars", type=int, default=160, help="Máximo de chars para considerar como título")

    args = p.parse_args(argv)

    src_dir = (SCRIPT_DIR / args.source_dir).resolve()
    out_dir = (SCRIPT_DIR / args.out_dir).resolve()

    if not src_dir.exists():
        print(f"[ERRO] source-dir não existe: {src_dir}")
        return 1

    out_dir.mkdir(parents=True, exist_ok=True)

    txt_files = iter_txt_files(src_dir)
    if not txt_files:
        print(f"[ERRO] nenhum .txt encontrado em: {src_dir}")
        return 1

    # monta params do S1
    cleaners = [c.strip() for c in args.cleaners.split(",") if c.strip()]
    lid_fallback = None if args.lid_fallback.strip().lower() == "none" else args.lid_fallback.strip()

    languages_supported = {x.strip() for x in args.languages.split(",") if x.strip()}
    if args.lang_hint:
        languages_supported = {args.lang_hint}

    params = S1Params(
        cleaners=cleaners,
        lid_primary=args.lid_primary,
        lid_fallback=lid_fallback,
        languages_supported=languages_supported,
        nlp_model_pref=args.nlp_size,
        paragraph_min_chars=args.paragraph_min_chars,
        lid_min_chars=args.lid_min_chars,
        # split_mode=args.split_mode,
        split_max_chars=args.split_max_chars,
        split_min_chars=args.split_min_chars,
        force_resegment_if_single=not args.no_resegment,
        force_title=args.force_title,
        title_scan_k=args.title_scan_k,
        title_max_chars=args.title_max_chars,       
    )

    ingestor = S1Ingestor(params)

    failures = 0
    batch_rows: List[dict] = []

    onto_path = (SCRIPT_DIR / args.onto).resolve()

    if not args.no_s2 and not onto_path.exists():
        print(f"[ERRO] ontologia não encontrada: {onto_path}")
        return 1

    # S3 só roda se S2 estiver ligado (dependência)
    do_s2_global = (not args.no_s2)
    do_s3_global = (not args.no_s3) and do_s2_global
    do_s4_global = (not args.no_s4) and do_s3_global  
    do_s5_global = (not args.no_s5) and do_s4_global     
    do_s6_global = (not args.no_s6) and do_s5_global
    do_s7_global = (not args.no_s7) and do_s6_global     

    for txt in txt_files:
        try:
            # process_one(txt, out_dir, ingestor)
            row = process_one(
                txt,
                out_dir,
                ingestor,
                do_s2=do_s2_global,
                do_s3=do_s3_global,
                onto_path=onto_path,
                base_ns=args.base_ns,
                s2_graph=args.s2_graph,
                s3_use_llm=(not args.s3_no_llm),
                s3_link_evokes=(not args.s3_no_evokes),
                do_s4=do_s4_global,
                s4_theta=args.theta,
                s4_w_keyword=args.w_keyword,
                s4_w_context=args.w_context,
                s4_w_strategy=args.w_strategy, 
                do_s5=do_s5_global,                
                s5_scope=args.s5_scope,
                s5_norm=args.s5_norm,
                s5_tau=args.s5_tau,
                s5_alpha_title=args.alpha_title,
                s5_alpha_body=args.alpha_body,
                s5_vec_places=args.s5_vec_places,
                s5_write_mi_vector_string=(not args.s5_no_mi_vector_string), 
                do_s6=do_s6_global,
                do_s7=do_s7_global,
                s7_inference=args.s7_inference,                                               
            )
            batch_rows.append(row)
        except Exception as e:
            failures += 1
            name = slugify(txt.stem)
            fail_dir = out_dir / name
            fail_dir.mkdir(parents=True, exist_ok=True)
            (fail_dir / "s1_error.txt").write_text(repr(e) + "\n", encoding="utf-8")
            print(f"[FALHOU] {txt.name}: {e!r}")

    # relatório geral do batch
    report = {
        "started_at": _iso_now(),
        "source_dir": str(src_dir),
        "out_dir": str(out_dir),
        "files_total": len(txt_files),
        "success": len(txt_files) - failures,
        "failures": failures,
        "docs": batch_rows,
    }
    _write_json(out_dir / "batch_report.json", report)

    # # Excel resumo do batch (1 linha por doc)
    # try:
    #     xlsx_path = write_batch_excel(out_dir, batch_rows)
    #     print(f"[XLSX OK] -> {xlsx_path.as_posix()}")
    # except Exception as e:
    #     print(f"[XLSX FALHOU] não consegui gerar batch_summary.xlsx: {e!r}")    

    print(f"\nConcluído. Sucesso: {len(txt_files) - failures} | Falhas: {failures} | Saída: {out_dir}")
    return 0 if failures == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
